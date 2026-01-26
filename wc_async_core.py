import csv
import io
from datetime import datetime
import pytz
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.platypus import KeepTogether
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart


def _format_folder_url(path: str | None) -> str:
    if not path:
        return "N/A"
    if isinstance(path, str) and path.startswith("gdrive://"):
        folder_id = path.split("gdrive://", 1)[-1]
        if folder_id:
            return f"https://drive.google.com/drive/folders/{folder_id}"
    return path


def generate_run_report_csv(stats: dict, results: list, run_config: dict) -> str:
    """Generate CSV report of extraction run with accurate cache stats"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header section
    writer.writerow(["Insolvency Document Extraction Report"])
    writer.writerow([f"Generated: {datetime.now(pytz.timezone('Europe/London')).strftime('%Y-%m-%d %H:%M:%S GMT')}"])
    writer.writerow([])
    
    # ✅ Extract and validate stats
    total_docs = stats.get('total_documents_downloaded', 0) or stats.get('total_documents_found', 0)
    fresh_docs = stats.get('total_documents_downloaded_fresh', 0) or stats.get('total_documents_fresh', 0)
    cached_docs = stats.get('documents_from_cache', 0)
    
    # ✅ VALIDATION: Ensure cache <= total
    if cached_docs > total_docs:
        total_docs = fresh_docs + cached_docs
    
    # Calculate metrics
    cache_rate = (cached_docs / max(1, total_docs)) * 100
    processing_time = stats.get('processing_time', 0)
    docs_per_min = (total_docs / max(0.016, processing_time / 60)) if processing_time > 0 else 0
    
    # Summary section
    writer.writerow(["SUMMARY"])
    writer.writerow(["Metric", "Value"])
    writer.writerow(["Date Range", f"{run_config.get('start_date', 'N/A')} to {run_config.get('end_date', 'N/A')}"])
    writer.writerow(["Categories", run_config.get('categories', 'N/A')])
    writer.writerow(["Companies Processed", stats.get('companies_processed', 0)])
    writer.writerow(["Successful", stats.get('companies_successful', 0)])
    writer.writerow(["Total Documents", total_docs])
    writer.writerow(["Fresh Downloads", fresh_docs])  # ✅ NEW
    writer.writerow(["From Cache", cached_docs])
    writer.writerow(["Cache Efficiency", f"{cache_rate:.1f}%"])
    writer.writerow(["Processing Time", f"{processing_time:.1f}s"])
    writer.writerow(["Documents per Minute", f"{docs_per_min:.0f}"])  # ✅ NEW
    writer.writerow([])
    
    # Details section
    writer.writerow(["COMPANY DETAILS"])
    writer.writerow([
        "Company Name",
        "Company Number", 
        "Procedure Type",
        "Documents Total",  # ✅ Changed from "Documents Downloaded"
        "Fresh Downloads",  # ✅ NEW
        "From Cache",
        "Status",
        "Folder Path"
    ])
    
    for r in results:
        # ✅ Validate individual company stats
        r_total = r.get('documents_downloaded', 0)
        r_fresh = r.get('documents_downloaded_fresh', 0)
        r_cache = r.get('documents_from_cache', 0)
        
        # Fix if cache > total
        if r_cache > r_total:
            r_total = r_fresh + r_cache
        
        writer.writerow([
            r.get('company_name', 'Unknown'),
            r.get('company_number', 'N/A'),
            r.get('notice_type', 'N/A'),
            r_total,
            r_fresh,  # ✅ NEW
            r_cache,
            '✅ Success' if r.get('success') else '❌ Failed',
            _format_folder_url(r.get('folder_path', 'N/A'))
        ])
    
    return output.getvalue()


def generate_run_report_excel(stats: dict, results: list, run_config: dict) -> bytes:
    """Generate formatted Excel report with multiple sheets and accurate cache stats"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None
    
    wb = openpyxl.Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header
    ws_summary['A1'] = "Insolvency Document Extraction Report"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A2'] = f"Generated: {datetime.now(pytz.timezone('Europe/London')).strftime('%Y-%m-%d %H:%M:%S GMT')}"
    
    # ✅ Extract and validate stats
    total_docs = stats.get('total_documents_downloaded', 0) or stats.get('total_documents_found', 0)
    fresh_docs = stats.get('total_documents_downloaded_fresh', 0) or stats.get('total_documents_fresh', 0)
    cached_docs = stats.get('documents_from_cache', 0)
    
    # ✅ VALIDATION: Ensure cache <= total
    if cached_docs > total_docs:
        total_docs = fresh_docs + cached_docs
    
    # Calculate metrics
    cache_rate = (cached_docs / max(1, total_docs)) * 100
    processing_time = stats.get('processing_time', 0)
    docs_per_min = (total_docs / max(0.016, processing_time / 60)) if processing_time > 0 else 0
    
    # Summary data
    summary_data = [
        ["Metric", "Value"],
        ["Date Range", f"{run_config.get('start_date', 'N/A')} to {run_config.get('end_date', 'N/A')}"],
        ["Categories", run_config.get('categories', 'N/A')],
        ["Companies Processed", stats.get('companies_processed', 0)],
        ["Successful", stats.get('companies_successful', 0)],
        ["Total Documents", total_docs],
        ["Fresh Downloads", fresh_docs],  # ✅ NEW
        ["From Cache", cached_docs],
        ["Cache Efficiency", f"{cache_rate:.1f}%"],
        ["Processing Time", f"{processing_time:.1f}s"],
        ["Documents per Minute", f"{docs_per_min:.0f}"],  # ✅ NEW
    ]
    
    for i, row in enumerate(summary_data, start=4):
        for j, value in enumerate(row, start=1):
            cell = ws_summary.cell(row=i, column=j, value=value)
            if i == 4:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
    
    # Auto-size columns
    for col in range(1, 3):
        ws_summary.column_dimensions[get_column_letter(col)].width = 25
    
    # Details sheet
    ws_details = wb.create_sheet("Company Details")
    headers = [
        "Company Name", 
        "Company Number", 
        "Procedure Type", 
        "Docs Total",  # ✅ Changed
        "Fresh",  # ✅ NEW
        "Cached", 
        "Status", 
        "Folder Path"
    ]
    
    for j, header in enumerate(headers, start=1):
        cell = ws_details.cell(row=1, column=j, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    for i, r in enumerate(results, start=2):
        # ✅ Validate individual company stats
        r_total = r.get('documents_downloaded', 0)
        r_fresh = r.get('documents_downloaded_fresh', 0)
        r_cache = r.get('documents_from_cache', 0)
        
        # Fix if cache > total
        if r_cache > r_total:
            r_total = r_fresh + r_cache
        
        ws_details.cell(row=i, column=1, value=r.get('company_name', 'Unknown'))
        ws_details.cell(row=i, column=2, value=r.get('company_number', 'N/A'))
        ws_details.cell(row=i, column=3, value=r.get('notice_type', 'N/A'))
        ws_details.cell(row=i, column=4, value=r_total)
        ws_details.cell(row=i, column=5, value=r_fresh)  # ✅ NEW
        ws_details.cell(row=i, column=6, value=r_cache)
        ws_details.cell(row=i, column=7, value='✅ Success' if r.get('success') else '❌ Failed')
        ws_details.cell(row=i, column=8, value=_format_folder_url(r.get('folder_path', 'N/A')))
    
    # Auto-size columns
    for col in range(1, 9):  # ✅ Updated to 9 columns
        ws_details.column_dimensions[get_column_letter(col)].width = 18
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def generate_run_report_pdf(stats: dict, results: list, run_config: dict) -> bytes:
    """Generate professional PDF report with charts and tables"""

    
    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e3a5f'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#2c5f8d'),
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    subheading_style = ParagraphStyle(
        'CustomSubHeading',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=colors.HexColor('#555555'),
        spaceAfter=8,
        fontName='Helvetica-Bold'
    )
    
    normal_style = styles['Normal']
    
    # Title
    story.append(Paragraph("Sarah Agent Extraction Report", title_style))
    
    # Timestamp
    uk_tz = pytz.timezone('Europe/London')
    timestamp = datetime.now(uk_tz).strftime('%d %B %Y, %H:%M:%S GMT')
    story.append(Paragraph(f"Generated: {timestamp}", normal_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Executive Summary Box
    story.append(Paragraph("Executive Summary", heading_style))

    # ✅ Extract and validate stats
    total_docs = stats.get('total_documents_downloaded', 0) or stats.get('total_documents_found', 0)
    fresh_docs = stats.get('total_documents_downloaded_fresh', 0) or stats.get('total_documents_fresh', 0)
    cached_docs = stats.get('documents_from_cache', 0)

    # ✅ VALIDATION: Ensure cache <= total
    if cached_docs > total_docs:
        total_docs = fresh_docs + cached_docs

    # Calculate metrics
    cache_rate = (cached_docs / max(1, total_docs)) * 100
    success_rate = (stats.get('companies_successful', 0) / max(1, stats.get('companies_processed', 1))) * 100
    processing_time = stats.get('processing_time', 0)
    docs_per_min = (total_docs / max(0.016, processing_time / 60)) if processing_time > 0 else 0

    summary_data = [
        ['Metric', 'Value'],
        ['Date Range', f"{run_config.get('start_date', 'N/A')} to {run_config.get('end_date', 'N/A')}"],
        ['Procedure Types', run_config.get('categories', 'All')],
        ['Companies Processed', str(stats.get('companies_processed', 0))],
        ['Successful Extractions', f"{stats.get('companies_successful', 0)} ({success_rate:.1f}%)"],
        ['Total Documents Retrieved', str(total_docs)],
        ['Fresh Downloads', str(fresh_docs)],  # ✅ NEW
        ['From Cache', f"{cached_docs} ({cache_rate:.1f}%)"],
        ['Processing Time', f"{processing_time:.1f} seconds"],
        ['Documents per Minute', f"{docs_per_min:.0f}"],  # ✅ NEW
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5f8d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Performance Charts
    story.append(Paragraph("Performance Metrics", heading_style))

    # Create success rate pie chart
    drawing = Drawing(400, 200)

    pie = Pie()
    pie.x = 50
    pie.y = 50
    pie.width = 150
    pie.height = 150

    successful = stats.get('companies_successful', 0)
    failed = stats.get('companies_processed', 0) - successful

    pie.data = [successful, failed] if successful + failed > 0 else [1, 0]
    pie.labels = [f'Successful\n{successful}', f'Failed\n{failed}'] if successful + failed > 0 else ['No Data', '']
    pie.slices[0].fillColor = colors.HexColor('#28a745')
    pie.slices[1].fillColor = colors.HexColor('#dc3545')

    drawing.add(pie)

    # ✅ Cache efficiency pie chart - FIXED to use fresh vs cache
    pie2 = Pie()
    pie2.x = 250
    pie2.y = 50
    pie2.width = 150
    pie2.height = 150

    # ✅ Use validated stats
    new_downloads = fresh_docs
    from_cache_chart = cached_docs

    pie2.data = [new_downloads, from_cache_chart] if (new_downloads + from_cache_chart) > 0 else [1, 0]
    pie2.labels = [f'Fresh\n{new_downloads}', f'Cached\n{from_cache_chart}'] if (new_downloads + from_cache_chart) > 0 else ['No Data', '']
    pie2.slices[0].fillColor = colors.HexColor('#007bff')
    pie2.slices[1].fillColor = colors.HexColor('#ffc107')

    drawing.add(pie2)

    story.append(drawing)
    story.append(Spacer(1, 0.3*inch))
    
    # Procedure Breakdown
    story.append(Paragraph("Breakdown by Procedure Type", heading_style))
    
    folder_structure = stats.get('folder_structure', {})
    if folder_structure:
        procedure_data = [['Procedure Type', 'Companies', 'Documents']]
        
        for folder, companies in folder_structure.items():
            company_count = len(companies)
            doc_count = sum(c.get('documents_downloaded', 0) for c in companies)
            procedure_data.append([folder, str(company_count), str(doc_count)])
        
        procedure_table = Table(procedure_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        procedure_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5f8d')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ]))
        
        story.append(procedure_table)
        story.append(Spacer(1, 0.2*inch))
    
    # Page break before detailed table
    story.append(PageBreak())
    
    # Detailed Company Results
    story.append(Paragraph("Detailed Company Results", heading_style))
    story.append(Spacer(1, 0.1*inch))

    # Table header - ✅ UPDATED
    detail_data = [['Company Name', 'Number', 'Type', 'Total', 'Fresh', 'Cache', 'Status']]

    # Add results (limit to reasonable number for PDF)
    for i, r in enumerate(results[:100]):  # Limit to first 100 companies
        status_symbol = '✓' if r.get('success') else '✗'
        
        # ✅ Validate individual company stats
        r_total = r.get('documents_downloaded', 0)
        r_fresh = r.get('documents_downloaded_fresh', 0)
        r_cache = r.get('documents_from_cache', 0)
        
        # Fix if cache > total
        if r_cache > r_total:
            r_total = r_fresh + r_cache
        
        detail_data.append([
            r.get('company_name', 'Unknown')[:40],  # Truncate long names
            r.get('company_number', 'N/A'),
            r.get('notice_type', 'N/A')[:20],
            str(r_total),  # ✅ Total
            str(r_fresh),  # ✅ NEW: Fresh
            str(r_cache),  # ✅ Cache
            status_symbol
        ])

    if len(results) > 100:
        detail_data.append(['...', f'{len(results) - 100} more companies', '...', '...', '...', '...', '...'])

    # Calculate optimal column widths - ✅ UPDATED for 7 columns
    col_widths = [2.3*inch, 0.9*inch, 1.1*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch]

    detail_table = Table(detail_data, colWidths=col_widths, repeatRows=1)
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5f8d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 0), (-1, -1), 'CENTER'),  # ✅ Align numbers center
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),  # ✅ Slightly smaller for more columns
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),  # ✅ Smaller for readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    story.append(detail_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Footer
    story.append(Spacer(1, 0.3*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    story.append(Paragraph(
        f"Report generated by Sarah Agent | {timestamp}",
        footer_style
    ))
    
    # Build PDF
    doc.build(story)
    
    buffer.seek(0)
    return buffer.getvalue()
